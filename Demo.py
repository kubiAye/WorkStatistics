import openpyxl

workBook = openpyxl.load_workbook("file/4月份-驻内考勤——考勤机-4月30日.xlsx")

print(workBook.sheetnames[0])


sheet = workBook.get_sheet_by_name(workBook.sheetnames[0])
rows = []
for row in sheet.iter_rows():
    # col_name = 2
    # col_date = 2
    # col_am = 2
    # col_pm = 2

    name = str(row[2].value)
    date = str(row[3].value)
    am = str(row[4].value)
    pm = str(row[6].value)
    # print(type(name))
    # print(type(date))
    # print(type(am))
    # print(type(pm))

    print(name + date + am + pm, end=" ")
    print("\r")
print(rows)
# for column in sheet[]:
#     for cell in column:
#         print()
